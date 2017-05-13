#encoding=utf-8

#创建一个表格，并设置格式、内容

from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Border,GradientFill
from openpyxl.styles import colors
from openpyxl import Workbook
from copy import copy

def create_table_and_data(file_path):
    wb = Workbook()
    ws = wb.active
    font = Font(name=u"微软雅黑",color=colors.RED,size=18,bold=True)
    thin = Side(border_style="thin", color=colors.BLACK)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    al = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="DDDDDD")
    ws.append([u"姓名", u"性别", u"学好",u"爱好",u"年龄"])
    for i in range(5):
        ws.append([u"张三", u"男", i,u"篮球",i*20])
    print ws.max_row
    print ws.max_column
    for cell in ws.rows[0]:
        cell.fill=fill
        cell.border=border
        cell.font=font
        cell.alignment=al
    font1 = Font(name=u"微软雅黑",color=colors.BLACK,size=12,bold=False)
    fill1 = PatternFill("solid", fgColor="DAADDD")
    for i in range(1,6):
        for cell in ws.rows[i]:
            cell.fill=fill1
            cell.border=border
            cell.font=font1
            cell.alignment=al

    wb.save(r"D:\test\excel\readfile.xlsx")